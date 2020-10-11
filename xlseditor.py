import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def wip_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet1 = wb["Sheet1"]
    for row in range(2, sheet1.max_row + 1):
        cell = sheet1.cell(row, 3)
        # this bit here is a price correction segment, we can change this to do whatever repetitive task we want to do
        updated_price = cell.value * .9
        updated_price_cell = sheet1.cell(row, 4)
        updated_price_cell = updated_price
        # im writing the updated data on a new cell, but irl we can add the updated data in a new workbook for data
        # security purposes
        values = Reference(sheet1,
                           min_row=2,
                           max_row=sheet1.max_row,
                           min_col=4,
                           max_col=4
                           )
    chart = BarChart()
    chart.add_data(values)
    sheet1.add_chart(chart, "a6")
    wb.save(filename)
